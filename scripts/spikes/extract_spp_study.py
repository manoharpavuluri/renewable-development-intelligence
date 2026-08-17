#!/usr/bin/env python3

import argparse
import hashlib
import json
import math
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def clean(value):
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        if not value:
            return None

        # SPP workbooks sometimes contain the literal string "None".
        # Treat that as absence, but preserve "N/A" because not-applicable
        # is semantically different from missing.
        if value.upper() == "NONE":
            return None

        return value

    return value


def text(value):
    value = clean(value)
    return "" if value is None else str(value)


def json_value(value):
    """
    Convert Excel/openpyxl values into JSON-safe values.
    """

    value = clean(value)

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None

    return value


def numeric(value):
    value = clean(value)

    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None


def sha256(path):
    digest = hashlib.sha256()

    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)

    return digest.hexdigest()


def make_unique_headers(values):
    """
    Normalize header whitespace while retaining recognizable
    source names.

    Duplicate headers receive __2, __3, etc.
    """

    headers = []
    seen = Counter()

    for index, value in enumerate(values, start=1):

        name = text(value)

        name = re.sub(r"\s+", " ", name).strip()

        if not name:
            name = f"_column_{index}"

        seen[name] += 1

        if seen[name] > 1:
            name = f"{name}__{seen[name]}"

        headers.append(name)

    return headers


def read_table(ws, header_row):
    headers = make_unique_headers(
        [
            cell.value
            for cell in ws[header_row]
        ]
    )

    rows = []

    for values in ws.iter_rows(
        min_row=header_row + 1,
        values_only=True,
    ):
        if not any(clean(v) is not None for v in values):
            continue

        row = {}

        for header, value in zip(headers, values):
            row[header] = json_value(value)

        rows.append(row)

    return rows


def contains(value, target):
    return target.upper() in text(value).upper()


def exact_rows(rows, field, target):
    target = text(target).upper()

    return [
        row
        for row in rows
        if text(row.get(field)).upper() == target
    ]


def prefix_rows(rows, field, target):
    target = text(target).upper()

    return [
        row
        for row in rows
        if text(row.get(field)).upper().startswith(target)
    ]


def unique_nonempty(values):
    result = []

    for value in values:
        value = clean(value)

        if value is None:
            continue

        if value not in result:
            result.append(value)

    return result


# ---------------------------------------------------------------------
# Workbook-specific extraction
# ---------------------------------------------------------------------

TABLE_DEFINITIONS = {
    "Requests": 2,
    "Seasonal LOIS": 1,
    "Constraints Summary": 1,
    "Assigned Upgrade Costs": 1,
    "Upgrade Summary": 1,
    "Contingent Upgrades - Screening": 1,
    "JTIQ Screening Summary": 1,
    "JTIQ Facility Screening Results": 1,
    "MISO Facility Screening Results": 1,
    "All Thermal": 1,
    "All Voltage": 1,
    "Stability Analysis Results": 1,
    "Short Circuit Analysis": 1,
    "SCRCCT Results": 1,
    "SCRCCT": 1,
}


def executive_summary(ws):
    values = []

    for row in ws.iter_rows(
        min_row=1,
        max_row=30,
        values_only=True,
    ):
        populated = [
            json_value(v)
            for v in row
            if clean(v) is not None
        ]

        if populated:
            values.append(populated)

    return values


def summarize_thermal(rows):
    thermal = [
        row
        for row in rows
        if text(row.get("SOLUTIONTYPE")).upper() == "THERMAL"
        and clean(row.get("MONTCOMMONNAME")) is not None
    ]

    maximum = None

    for row in thermal:
        loading = numeric(row.get("TC%LOADING"))

        if loading is None:
            continue

        if maximum is None or loading > maximum["loading_pct"]:
            maximum = {
                "loading_pct": loading,
                "facility": row.get("MONTCOMMONNAME"),
                "season": row.get("SEASON"),
                "contingency": row.get("CONTNAME"),
                "upgrade": row.get("Upgrade Name"),
            }

    return {
        "result_count": len(thermal),
        "max_transfer_case_loading": maximum,
        "facilities": unique_nonempty(
            row.get("MONTCOMMONNAME")
            for row in thermal
        ),
        "upgrades": unique_nonempty(
            row.get("Upgrade Name")
            for row in thermal
        ),
    }


def summarize_voltage(rows):
    actual = [
        row
        for row in rows
        if clean(row.get("MONTCOMMONNAME")) is not None
    ]

    return {
        "result_count": len(actual),
        "facilities": unique_nonempty(
            row.get("MONTCOMMONNAME")
            for row in actual
        ),
    }


def summarize_stability(rows):
    criteria = [
        "Rotor Angle Stability",
        "Transient Voltage Response > 0.7 p.u.",
        "Transient Voltage Response < 1.2 p.u.",
        "Post Fault Steady State Voltage > 0.9 p.u.",
        "Post Fault Steady State Voltage < 1.1 p.u.",
        "Damping Factor > 0.8 %",
        "Low Voltage Rides Through",
    ]

    failures = {}

    for criterion in criteria:
        count = sum(
            1
            for row in rows
            if text(row.get(criterion)).upper() == "NO"
        )

        failures[criterion] = count

    violation_counter = Counter(
        text(row.get("Violation Summary"))
        for row in rows
        if clean(row.get("Violation Summary")) is not None
    )

    mitigation_counter = Counter(
        text(row.get("Primary Mitigation"))
        for row in rows
        if clean(row.get("Primary Mitigation")) is not None
    )

    return {
        "events_analyzed": len(rows),
        "raw_criterion_no_counts": failures,
        "unique_violation_patterns": len(violation_counter),
        "top_violation_patterns": [
            {
                "count": count,
                "text": value,
            }
            for value, count
            in violation_counter.most_common(10)
        ],
        "top_primary_mitigations": [
            {
                "count": count,
                "text": value,
            }
            for value, count
            in mitigation_counter.most_common(10)
        ],
    }


def summarize_short_circuit(rows):
    changes = [
        numeric(row.get("Change in Fault Current (kA)"))
        for row in rows
    ]

    changes = [
        value
        for value in changes
        if value is not None
    ]

    breaker_issues = [
        row
        for row in rows
        if clean(row.get("Circuit Breakers Exceeding Capacity"))
        not in (None, "None", "N/A")
    ]

    required_facilities = [
        row
        for row in rows
        if clean(row.get("Facilities Required to Interconnect"))
        not in (None, "None", "N/A")
    ]

    return {
        "buses_analyzed": len(rows),
        "maximum_fault_current_change_ka": (
            max(changes)
            if changes
            else None
        ),
        "breaker_capacity_issue_rows": len(breaker_issues),
        "required_facility_rows": len(required_facilities),
    }


def summarize_scrcct(rows):
    scr_values = [
        numeric(row.get("SCR"))
        for row in rows
    ]

    scr_values = [
        value
        for value in scr_values
        if value is not None
    ]

    cct_values = [
        numeric(row.get("CCT"))
        for row in rows
    ]

    cct_values = [
        value
        for value in cct_values
        if value is not None
    ]

    return {
        "cases_analyzed": len(rows),

        "minimum_scr": (
            min(scr_values)
            if scr_values
            else None
        ),

        "scr_failures": sum(
            1
            for row in rows
            if text(row.get("SCR PASS/FAIL")).upper() == "FAIL"
        ),

        "minimum_cct": (
            min(cct_values)
            if cct_values
            else None
        ),

        "cct_failures": sum(
            1
            for row in rows
            if text(row.get("CCT PASS/FAIL")).upper() == "FAIL"
        ),
    }


def summarize_upgrades(rows):
    known_total_cost = 0.0
    known_count = 0
    tbd_count = 0

    upgrades = []

    for row in rows:

        cost = numeric(
            row.get("Total Upgrade Cost")
        )

        if cost is not None:
            known_total_cost += cost
            known_count += 1
        else:
            tbd_count += 1

        upgrades.append(
            {
                "type": row.get("Upgrade Type"),
                "name": row.get("Upgrade Name"),
                "details": row.get("Upgrade Details"),
                "transmission_owner": row.get(
                    "Transmission Owner(s)"
                ),
                "estimated_lead_time": row.get(
                    "Estimated Lead Time"
                ),
                "total_upgrade_cost": row.get(
                    "Total Upgrade Cost"
                ),
            }
        )

    return {
        "upgrade_count": len(rows),
        "known_cost_upgrade_count": known_count,
        "tbd_cost_upgrade_count": tbd_count,
        "known_total_upgrade_cost": known_total_cost,
        "upgrades": upgrades,
    }


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()

    parser.add_argument(
        "workbook",
        type=Path,
    )

    parser.add_argument(
        "--request-id",
        required=True,
    )

    parser.add_argument(
        "--output",
        type=Path,
        default=None,
    )

    args = parser.parse_args()

    path = args.workbook
    request_id = args.request_id

    if not path.exists():
        raise SystemExit(
            f"Workbook not found: {path}"
        )

    wb = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    missing_sheets = [
        sheet
        for sheet in TABLE_DEFINITIONS
        if sheet not in wb.sheetnames
    ]

    if missing_sheets:
        print(
            "WARNING: Missing expected sheets:",
            ", ".join(missing_sheets),
        )

    tables = {}

    for sheet_name, header_row in TABLE_DEFINITIONS.items():

        if sheet_name not in wb.sheetnames:
            tables[sheet_name] = []
            continue

        tables[sheet_name] = read_table(
            wb[sheet_name],
            header_row,
        )

    # --------------------------------------------------------------
    # Project-specific filtering
    # --------------------------------------------------------------

    requests = exact_rows(
        tables["Requests"],
        "Gen Number",
        request_id,
    )

    seasonal_lois = exact_rows(
        tables["Seasonal LOIS"],
        "Gen Number",
        request_id,
    )

    constraints = prefix_rows(
        tables["Constraints Summary"],
        "Study Name",
        request_id,
    )

    assigned_costs = exact_rows(
        tables["Assigned Upgrade Costs"],
        "Gen Number",
        request_id,
    )

    upgrades = prefix_rows(
        tables["Upgrade Summary"],
        "Study Name",
        request_id,
    )

    contingent = exact_rows(
        tables["Contingent Upgrades - Screening"],
        "Gen Number",
        request_id,
    )

    jtiq = exact_rows(
        tables["JTIQ Screening Summary"],
        "Gen Number",
        request_id,
    )

    thermal = exact_rows(
        tables["All Thermal"],
        "SOURCE",
        request_id,
    )

    voltage = prefix_rows(
        tables["All Voltage"],
        "Study Name",
        request_id,
    )

    stability = prefix_rows(
        tables["Stability Analysis Results"],
        "Study Name",
        request_id,
    )

    short_circuit = exact_rows(
        tables["Short Circuit Analysis"],
        "Gen Number",
        request_id,
    )

    scrcct_results = exact_rows(
        tables["SCRCCT Results"],
        "GI Number",
        request_id,
    )

    scrcct = exact_rows(
        tables["SCRCCT"],
        "Gen Number",
        request_id,
    )

    # --------------------------------------------------------------
    # Assigned-cost distinction
    # --------------------------------------------------------------

    allocated_costs = [
        numeric(row.get("Allocated Cost"))
        for row in assigned_costs
    ]

    allocated_costs = [
        value
        for value in allocated_costs
        if value is not None
    ]

    # --------------------------------------------------------------
    # Output
    # --------------------------------------------------------------

    result = {
        "source": {
            "authority": "Southwest Power Pool",
            "artifact_name": path.name,
            "sha256": sha256(path),
            "request_id": request_id,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
        },

        "executive_summary": (
            executive_summary(
                wb["Executive Summary"]
            )
            if "Executive Summary" in wb.sheetnames
            else []
        ),

        "request": (
            requests[0]
            if requests
            else None
        ),

        "seasonal_lois": seasonal_lois,

        "constraints": constraints,

        "assigned_upgrade_costs": assigned_costs,

        "assigned_cost_summary": {
            "known_allocated_cost_total": sum(
                allocated_costs
            ),
            "contains_tbd_allocated_cost": any(
                numeric(row.get("Allocated Cost")) is None
                and clean(row.get("Allocated Cost")) is not None
                for row in assigned_costs
            ),
        },

        "upgrade_summary": summarize_upgrades(
            upgrades
        ),

        "contingent_upgrades": contingent,

        "jtiq_screening": jtiq,

        "thermal_summary": summarize_thermal(
            thermal
        ),

        "voltage_summary": summarize_voltage(
            voltage
        ),

        "stability_summary": summarize_stability(
            stability
        ),

        "short_circuit_summary": summarize_short_circuit(
            short_circuit
        ),

        "scrcct_results": scrcct_results,

        "scrcct_summary": summarize_scrcct(
            scrcct
        ),

        "row_counts": {
            "requests": len(requests),
            "seasonal_lois": len(seasonal_lois),
            "constraints": len(constraints),
            "assigned_upgrade_costs": len(assigned_costs),
            "upgrade_summary": len(upgrades),
            "contingent_upgrades": len(contingent),
            "thermal": len(thermal),
            "voltage": len(voltage),
            "stability": len(stability),
            "short_circuit": len(short_circuit),
            "scrcct": len(scrcct),
        },
    }

    if args.output is None:
        output_path = (
            path.parent
            / f"{request_id}_extracted.json"
        )
    else:
        output_path = args.output

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_path.write_text(
        json.dumps(
            result,
            indent=2,
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    print("=== SPP STUDY EXTRACTION ===")
    print("Request:", request_id)
    print("Workbook:", path)
    print("SHA256:", result["source"]["sha256"])

    print("\n=== REQUEST ===")

    if result["request"]:
        print(
            "MW:",
            result["request"].get("MW Amount"),
        )
        print(
            "Fuel:",
            result["request"].get("Fuel Type"),
        )
        print(
            "POI:",
            result["request"].get("POI"),
        )
        print(
            "Service:",
            result["request"].get("Service"),
        )

    print("\n=== CONSTRAINTS ===")
    print("Rows:", len(constraints))

    for row in constraints:
        constraint_type = row.get("Constraint Type")
        constraint = row.get("Constraints")
        upgrade = row.get("Upgrade Name")

        if (
            clean(constraint) is not None
            or clean(upgrade) is not None
        ):
            print(
                f"- {constraint_type}: "
                f"{constraint} -> {upgrade}"
            )

    print("\n=== COST DISTINCTION ===")
    print(
        "Known allocated cost:",
        f"${result['assigned_cost_summary']['known_allocated_cost_total']:,.2f}",
    )
    print(
        "Known total upgrade cost represented in study:",
        f"${result['upgrade_summary']['known_total_upgrade_cost']:,.2f}",
    )
    print(
        "Upgrade rows with TBD cost:",
        result["upgrade_summary"]["tbd_cost_upgrade_count"],
    )

    print("\n=== THERMAL ===")
    print(
        json.dumps(
            result["thermal_summary"],
            indent=2,
        )
    )

    print("\n=== STABILITY ===")
    print(
        "Events analyzed:",
        result["stability_summary"]["events_analyzed"],
    )
    print(
        "Raw criterion NO counts:",
        result["stability_summary"][
            "raw_criterion_no_counts"
        ],
    )

    print("\n=== SHORT CIRCUIT ===")
    print(
        json.dumps(
            result["short_circuit_summary"],
            indent=2,
        )
    )

    print("\n=== SCR/CCT ===")
    print(
        json.dumps(
            result["scrcct_summary"],
            indent=2,
        )
    )

    print("\nOutput:")
    print(output_path)


if __name__ == "__main__":
    main()
